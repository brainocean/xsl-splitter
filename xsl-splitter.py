#!/usr/bin/env python3

import os
from itertools import groupby
from openpyxl import Workbook, load_workbook
import PySimpleGUI as sg

def load_source_data(filename):
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    data = []
    for row in sheet.values:
        rowdata = []
        for value in row:
            rowdata.append(value)
        data.append(rowdata)
    return data

def get_col_index(line, pred):
    for i in range(len(line)):
        if line[i] and pred(line[i]):
            return i
    raise Exception("找不到指定的列")

def group_data(source_data, col_name):
    unit_name_col = get_col_index(source_data[0], lambda line:col_name in line)
    keyfunc = lambda r:r[unit_name_col]
    grouped_data = groupby(sorted(source_data[1:], key=keyfunc), keyfunc)
    target_data = {}
    for key, rows_gen in grouped_data:
        rows = []
        for row in rows_gen:
            rows.append(row)
        target_data[key] = rows
    return target_data

def write_one_target_xls(filename, header_line, rows):
    wb = Workbook()
    s = wb.active
    s.append(header_line)
    for r in rows:
        s.append(r)
    wb.save(filename=filename)

def write_target_xls(target_folder, source_file, header_line, data):
    folder_name = os.path.join(target_folder, os.path.splitext(source_file)[0])
    if not os.path.isdir(folder_name):
        os.makedirs(folder_name)
    for k in data.keys():
        print("正在写入", k, len(data[k]), "行数据")
        target_filename = os.path.join(folder_name, k+'.xlsx')
        write_one_target_xls(target_filename, header_line, data[k])

source_data = None

sg.theme('Dark Grey 13')

def updateColumnCombo(window, cols):
    window['-column-combo-'].update(values=cols)

layout = [[sg.Text('源数据文件')],
          [sg.Input(key='-source-filename-', enable_events=True, expand_x=True, disabled=True),
           sg.FileBrowse(button_text='打开', target='-source-filename-', file_types=(('Excel', '*.xls *.xlsx'),))],
          [sg.Text('选择用来分组的列')],
          [sg.Combo([], key='-column-combo-', expand_x=True, enable_events=True, readonly=True)],
          [sg.Output(expand_x=True, s=(20,30))],
          [sg.Text('目标文件夹')],
          [sg.Input(key='-target-folder-', expand_x=True, disabled=True),
           sg.FolderBrowse(button_text='浏览')],
          [sg.Button('执行', disabled=True), sg.Button('退出')],
          ]

window = sg.Window('数据文件分割器', layout, finalize=True)

while True:  # Event Loop
    event, values = window.read()
    if event == "-source-filename-":
        source_path = values['-source-filename-']
        print("读入" + source_path)
        source_data = load_source_data(source_path)
        updateColumnCombo(window, source_data[0])
        window["-target-folder-"].update(os.path.join(os.path.dirname(source_path), "target"))
        continue
    if event == '-column-combo-':
        if values['-column-combo-'] != None:
            window['执行'].update(disabled=False)
        continue
    if event == sg.WIN_CLOSED or event == '退出':
        break
    if event == '执行':
        print("分割后文件写入文件夹", values['-target-folder-'])
        write_target_xls(values['-target-folder-'],
                         os.path.basename(values['-source-filename-']),
                         source_data[0],
                         group_data(source_data, values['-column-combo-']))
        print("文件" + values['-source-filename-'] + " 分割完成.")

window.close()
